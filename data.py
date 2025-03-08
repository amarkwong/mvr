import pandas as pd
import os
import re
import json
from ui import styled_print
from openpyxl import load_workbook

# * Utilitiy functions
def detect_percentage_format(file_path, sheet_name=None):
    """
    Detects which columns in an Excel sheet have percentage formatting.

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str, optional): Sheet to inspect. Defaults to the first sheet.

    Returns:
        dict: Mapping of column names to a boolean indicating if they are formatted as percentages.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active  # Use the first sheet if none specified

    percentage_columns = {}

    for col in ws.iter_cols():
        col_letter = col[0].column_letter
        header = col[0].value  # Assuming first row contains column headers
        if not header:
            continue

        # Check if any cell in this column has a percentage format
        has_percentage = any(
            cell.number_format in ["0%", "0.00%", "0.0%", "0.000%"] for cell in col if cell.value is not None
        )

        if has_percentage:
            percentage_columns[header] = True

    return percentage_columns


def standardize_numeric_columns(df, percentage_columns):
    """
    Standardizes numeric columns by:
      - Converting string values (e.g., "90%") into numbers.
      - Ensuring consistent formatting across all rows.

    Parameters:
        df (pd.DataFrame): The dataset.
        percentage_columns (dict): Dictionary mapping column names to whether they are percentages.

    Returns:
        pd.DataFrame: The modified DataFrame with standardized numeric formats.
    """
    for col, is_percentage in percentage_columns.items():
        if col in df.columns:
            # ✅ Convert any percentage strings like "90%" to numeric values
            df[col] = df[col].astype(str).str.replace("%", "").str.strip()
            df[col] = pd.to_numeric(df[col], errors="coerce")  # Convert to numbers safely

            # ✅ Identify inconsistent values
            inconsistent_values = (df[col] > 1).sum() > 0  # If values > 1, they are likely whole numbers

            if is_percentage and inconsistent_values:
                print(f"📌 Converting {col} from percentage to decimal.")
                df[col] = df[col] / 100  # Convert whole numbers (e.g., 90 → 0.9)

            elif not is_percentage and (df[col] < 1).sum() > 0:
                print(f"📌 Converting {col} from decimal to percentage scale.")
                df[col] = df[col] * 100  # Convert decimals (e.g., 0.9 → 90)

    return df

def process_headers(columns):
    header_names = []
    descriptions = {}
    for col in columns:
        col_str = str(col).strip()
        if '#' in col_str:
            actual, desc = col_str.split('#', 1)
            actual = actual.strip().lstrip('*').rstrip('?')
            desc = desc.strip()
            header_names.append(actual)
            descriptions[actual] = desc
        else:
            header_names.append(col_str.lstrip('*'))
    return header_names, descriptions

def combine_gene_info(row):
    """
    Combine gene-related columns into a nested list of dictionaries.
    Each gene record is taken from the corresponding entries in the aggregated lists.
    Skip any gene record where the gene value is missing or equals 0 (or "0").
    """
    gene_records = []
    genes = row["Gene"]
    vafs = row["VAF% G1"]
    tiers = row["Tier"]
    var_descs = row["Variant description"]
    for gene, vaf, tier, var_desc in zip(genes, vafs, tiers, var_descs):
        if pd.isna(gene) or gene == 0 or str(gene).strip() == "0":
            continue
        gene_records.append({
            "name": gene,
            "VAF% G1": vaf,
            "Tier": tier,
            "Variant description": var_desc
        })
    return gene_records

def generate_metadata_mapping(header_metadata):
    """
    Converts the metadata dictionary into a lookup format where numeric values are mapped to labels.

    Parameters:
        header_metadata (dict): Dictionary with column metadata mappings.

    Returns:
        dict: Reformatted dictionary where numeric values are mapped to readable labels.
    """
    metadata_lookup = {}

    for column, mapping in header_metadata.items():
        if isinstance(mapping, dict):  # ✅ Fix: Handle dictionary mappings
            metadata_lookup[column] = mapping  # Directly store if it's a dict
        elif isinstance(mapping, str):  # ✅ Handle string-based mappings
            matches = re.findall(r"(\d+)\s*[=:-]?\s*([\w\s]+)", mapping)  
            metadata_lookup[column] = {int(num): label for num, label in matches}

    return metadata_lookup

def data_cleansing(file_path):
    """
    Load Excel data, aggregate rows by 'UR', and ensure consistent numeric formats.

    Returns:
      - aggregated: the cleaned, aggregated DataFrame.
      - header_metadata: dictionary of header descriptions.
    """
    na_values = ["NA", "na", "N/A", "n/a", "N/a"]
    df = pd.read_excel(file_path, engine="openpyxl", na_values=na_values)
    df["UR"] = df["UR"].ffill()

    # Detect numeric columns with mixed formats
    percentage_columns = detect_percentage_format(file_path)

    # Standardize numeric formats
    df = standardize_numeric_columns(df, percentage_columns)

    # Process headers
    raw_headers = list(df.columns)
    new_columns, header_metadata = process_headers(raw_headers)
    df.columns = new_columns

    # Sort by patient identifier (UR) to ensure rows for the same patient remain together
    if "UR" in df.columns:
        df = df.sort_values("UR")

    # Define gene-related columns
    gene_columns = ["Gene", "VAF% G1", "Tier", "Variant description"]

    # Build an aggregation dictionary
    agg_dict = {}
    for col in df.columns:
        if col == "UR":
            continue  # Grouping key
        if col in gene_columns:
            agg_dict[col] = lambda x, col=col: x.tolist()
        else:
            agg_dict[col] = lambda x: x.dropna().iloc[0] if not x.dropna().empty else None

    aggregated = df.groupby("UR", as_index=False).agg(agg_dict)

    # Combine gene-related columns into a single nested "Gene" list
    aggregated["Gene"] = aggregated.apply(combine_gene_info, axis=1)
    aggregated = aggregated.drop(columns=["VAF% G1", "Tier", "Variant description"])

    aggregated.to_json("merged_output.json", orient="records", indent=2)

    return aggregated, header_metadata

def data_derive(df, config_path="config.json"):
    """
    Derives new columns based on rules specified in config.json.

    - Supports Count (for lists)
    - Supports Mapping (numeric ranges to categories)
    - Supports Conditional Logic (multi-column conditions)

    Parameters:
        df (pd.DataFrame): The dataset.
        config_path (str): Path to the config file.

    Returns:
        pd.DataFrame: Updated dataset with derived columns.
        dict: Updated metadata mapping for categorical labels.
    """
    # ✅ Load configuration
    with open(config_path, "r") as f:
        config = json.load(f)

    derivation_config = config.get("data", {}).get("data_derivation", {})
    columns_to_derive = derivation_config.get("columns", [])

    header_metadata = {}  # Stores category labels

    for col_def in columns_to_derive:
        col_name = col_def["name"]
        method = col_def["method"]

        if method == "Count":
            base_col = col_def["base_column"]
            if base_col in df.columns:
                df[col_name] = df[base_col].apply(lambda x: len(x) if isinstance(x, list) else 0)

        elif method == "mapping":
            base_col = col_def["base_column"]
            mapping_rules = col_def.get("map", [])

            def map_value(value):
                for rule in mapping_rules:
                    category, criteria = list(rule.items())[0]
                    floor = criteria.get("floor", float("-inf"))
                    ceiling = criteria.get("ceiling", float("inf"))
                    int_value = criteria["int_value"]

                    if floor <= value < ceiling:
                        return int_value
                return None  # No match

            if base_col in df.columns:
                df[col_name] = df[base_col].apply(lambda x: map_value(x) if pd.notna(x) else None)

            # ✅ Store label mapping
            header_metadata[col_name] = {v["int_value"]: k for rule in mapping_rules for k, v in rule.items()}

        elif method == "conditional":
            conditions = col_def.get("conditions", [])
            default_value = col_def.get("default", {}).get("value", None)
            default_label = col_def.get("default", {}).get("label", "Unknown")

            def evaluate_conditions(row):
                for cond in conditions:
                    if_clause = cond.get("if", {})
                    then_clause = cond.get("then", {})

                    # ✅ Check if all conditions are met
                    if all(
                        (
                            (row[col] > crit["greater_than"]) if "greater_than" in crit else True and
                            (row[col] < crit["less_than"]) if "less_than" in crit else True
                        )
                        for col, crit in if_clause.items()
                    ):
                        return then_clause["value"]

                return default_value  # Return default if no conditions match

            # ✅ Apply conditions to dataset
            df[col_name] = df.apply(evaluate_conditions, axis=1)

            # ✅ Store labels in metadata
            header_metadata[col_name] = {cond["then"]["value"]: cond["then"]["label"] for cond in conditions}
            header_metadata[col_name][default_value] = default_label

    return df, header_metadata

def data_fitting(df, columns_to_check, config_path="config.json"):
    """
    Handles missing data based on predefined rules in the configuration file.

    Parameters:
        df (pd.DataFrame): The DataFrame containing the data.
        columns_to_check (list): List of columns to check for missing values.
        config_path (str): Path to the main config file.

    Returns:
        pd.DataFrame: Updated DataFrame with missing values handled.
    """
    allowed_options = ['drop', 'mean', 'median', 'mode', 'zero', 'calc']

    # Load config
    if os.path.exists(config_path):
        with open(config_path, "r") as f:
            config = json.load(f)
    else:
        config = {"data": {"data_fitting": {}}}

    # Get missing data handling rules
    data_fitting_config = config.get("data", {}).get("data_fitting", {})

    for col in columns_to_check:
        if col not in df.columns:
            print(f"⚠️ Column '{col}' not found in DataFrame, skipping.")
            continue

        # Get missing rows
        missing_urs = df.loc[df[col].isnull(), "UR"].tolist()
        if missing_urs:
            styled_print(f"⚠️ Missing values found in '{col}': {missing_urs}")

            # Get the pre-configured option
            option = data_fitting_config.get(col, None)

            if not option:
                styled_print(f"❓ Choose a method to handle missing data for '{col}': ['drop', 'mean', 'median', 'mode', 'zero', 'calc']")
                option = input("Enter choice: ").strip().lower()
                while option not in allowed_options:
                    option = input("Invalid choice. Please enter: 'drop', 'mean', 'median', 'mode', 'zero', 'calc': ").strip().lower()

                data_fitting_config[col] = option
                config["data"]["data_fitting"] = data_fitting_config
                with open(config_path, "w") as f:
                    json.dump(config, f, indent=2)

            # Apply the chosen option
            if isinstance(option, dict) or option == "calc":
                calc_conf = option if isinstance(option, dict) else None
                if calc_conf is None:
                    print(f"⚠️ Calculation configuration for '{col}' is missing, skipping calculation.")
                else:
                    first_input = calc_conf["first_input"]
                    second_input = calc_conf["second_input"]
                    op = calc_conf["operator"]
                    unit = calc_conf["unit"].lower()

                    if op != "-" or unit != "month":
                        print(f"⚠️ Only '-' and 'month' are supported for '{col}', skipping calculation.")
                    else:
                        df[first_input] = pd.to_datetime(df[first_input], errors="coerce")
                        df[second_input] = pd.to_datetime(df[second_input], errors="coerce")

                        df.loc[:, col] = df.apply(
                            lambda row: (row[first_input] - row[second_input]).days / 30 if pd.isnull(row[col]) and pd.notnull(row[first_input]) and pd.notnull(row[second_input]) else row[col],
                            axis=1
                        )
                        styled_print(f"✅ '{col}' values calculated using: {first_input} {op} {second_input} in {unit}s.")

            elif option == "drop":
                df = df.dropna(subset=[col])
                styled_print(f"✅ Rows with missing '{col}' have been dropped.")

            elif option == "mean":
                if pd.api.types.is_numeric_dtype(df[col]):
                    df.loc[:, col] = df[col].fillna(df[col].mean())  # ✅ Best practice: No inplace=True
                    styled_print(f"✅ Missing '{col}' values filled with mean.")
                else:
                    print(f"⚠️ '{col}' is not numeric; skipping mean imputation.")

            elif option == "median":
                if pd.api.types.is_numeric_dtype(df[col]):
                    df.loc[:, col] = df[col].fillna(df[col].median())  # ✅ No inplace=True
                    styled_print(f"✅ Missing '{col}' values filled with median.")
                else:
                    print(f"U+26A0'{col}' is not numeric; skipping median imputation.")

            elif option == "mode":
                mode_val = df[col].mode()
                if not mode_val.empty:
                    df.loc[:, col] = df[col].fillna(mode_val.iloc[0])  # ✅ No inplace=True
                    styled_print(f"✅ Missing '{col}' values filled with mode.")
                else:
                    print(f"⚠️ '{col}' has no mode; skipping mode imputation.")

            elif option == "zero":
                df.loc[:, col] = df[col].fillna(0)  # ✅ No inplace=True
                styled_print(f"✅ Missing '{col}' values filled with zero.")

        else:
            print(f"✅ No missing data detected in '{col}'.")

    return df